from flask_restful import Resource
from flask_jwt_extended import jwt_required
from flask import jsonify, Response, request, send_file
from werkzeug.utils import secure_filename
from app.modelos.modelos import Libros, db, fecha_modificacion, EstadisticasPorMes, EstadisticasPorMesAuxiliar
from openpyxl import Workbook, load_workbook
from io import StringIO, BytesIO
import csv
from datetime import datetime


def ClonarEstadisticas():
    db.session.query(EstadisticasPorMes).delete()
    db.create_all()
    db.session.query(EstadisticasPorMesAuxiliar).delete()
    db.session.commit()
    estadisticas =EstadisticasPorMes.query.all()
    for estadistica in estadisticas:
        nuevaEstadistica = EstadisticasPorMesAuxiliar(
            mes = estadistica.mes,
            anyo = estadistica.anyo,
            numero_libros = estadistica.numero_libros,
            numero_estimaciones = estadistica.numero_estimaciones,
            numero_usuarios = estadistica.numero_usuarios,
            libro_mas_visitado = estadistica.libro_mas_visitado,
            isbn_libro_mas_visitado = estadistica.isbn_libro_mas_visitado,
            visitas_libro_mas_visitado = estadistica.visitas_libro_mas_visitado,
            url_imagen_libro_mas_visitado = estadistica.url_imagen_libro_mas_visitado,
            numero_visitas_totales = estadistica.numero_visitas_totales,
            numero_sugerencias = estadistica.numero_sugerencias
        )
        db.session.add(nuevaEstadistica)
    db.session.query(EstadisticasPorMes).delete()
    db.session.commit()

def  ActualizarEstadisticasImportacion(libro : Libros):
    estadistica_existente = EstadisticasPorMesAuxiliar.query.filter_by(mes=libro.mes_creacion, anyo=libro.anyo_creacion).first()
    estadistica_nueva = EstadisticasPorMes.query.filter_by(mes=libro.mes_creacion, anyo=libro.anyo_creacion).first()

    if int(libro.mes_creacion) == 1:
        mes_anterior = 12
        anyo_anterior = int(libro.anyo_creacion) - 1
        estadistica_anterior = EstadisticasPorMesAuxiliar.query.filter_by(mes=mes_anterior, anyo=anyo_anterior).first()
    else:
        mes_anterior = int(libro.mes_creacion) - 1
        anyo_anterior = int(libro.anyo_creacion)
        estadistica_anterior = EstadisticasPorMesAuxiliar.query.filter_by(mes=mes_anterior, anyo=anyo_anterior).first()
    libros_totales = estadistica_anterior.numero_libros if estadistica_anterior else 0
    
    
    if estadistica_existente:
        estadistica_existente.numero_libros = estadistica_existente.numero_libros + 1
        estadistica_existente.numero_visitas_totales = estadistica_existente.numero_visitas_totales + int(libro.visitas_mensuales)
        if estadistica_existente.visitas_libro_mas_visitado < int(libro.visitas_mensuales):
            estadistica_existente.libro_mas_visitado = libro.titulo
            estadistica_existente.isbn_libro_mas_visitado = libro.isbn
            estadistica_existente.visitas_libro_mas_visitado = int(libro.visitas_mensuales)
            estadistica_existente.url_imagen_libro_mas_visitado = libro.url_imagen
        if estadistica_nueva:
            estadistica_nueva.mes = estadistica_existente.mes
            estadistica_nueva.anyo = estadistica_existente.anyo
            estadistica_nueva.numero_libros = estadistica_existente.numero_libros
            estadistica_nueva.numero_visitas_totales = estadistica_existente.numero_visitas_totales
            estadistica_nueva.numero_estimaciones = estadistica_existente.numero_estimaciones
            estadistica_nueva.numero_usuarios = estadistica_existente.numero_usuarios
            estadistica_nueva.libro_mas_visitado = estadistica_existente.libro_mas_visitado
            estadistica_nueva.isbn_libro_mas_visitado = estadistica_existente.isbn_libro_mas_visitado
            estadistica_nueva.visitas_libro_mas_visitado = estadistica_existente.visitas_libro_mas_visitado
            estadistica_nueva.url_imagen_libro_mas_visitado = estadistica_existente.url_imagen_libro_mas_visitado
            estadistica_nueva.numero_sugerencias = estadistica_existente.numero_sugerencias
        else:
            nuevaEstadistica = EstadisticasPorMes(
            mes = estadistica_existente.mes,
            anyo = estadistica_existente.anyo,
            numero_libros = libros_totales + 1,
            numero_estimaciones = estadistica_existente.numero_estimaciones,
            numero_usuarios = estadistica_existente.numero_usuarios,
            libro_mas_visitado = estadistica_existente.libro_mas_visitado,
            isbn_libro_mas_visitado = estadistica_existente.isbn_libro_mas_visitado,
            visitas_libro_mas_visitado = estadistica_existente.visitas_libro_mas_visitado,
            url_imagen_libro_mas_visitado = estadistica_existente.url_imagen_libro_mas_visitado,
            numero_visitas_totales = estadistica_existente.numero_visitas_totales,
            numero_sugerencias = estadistica_existente.numero_sugerencias
            )
            db.session.add(nuevaEstadistica)
    else :
        if estadistica_nueva:
            if estadistica_nueva.visitas_libro_mas_visitado < int(libro.visitas_mensuales):
                estadistica_nueva.libro_mas_visitado = libro.titulo
                estadistica_nueva.isbn_libro_mas_visitado = libro.isbn
                estadistica_nueva.visitas_libro_mas_visitado = int(libro.visitas_mensuales)
                estadistica_nueva.url_imagen_libro_mas_visitado = libro.url_imagen
            
            estadistica_nueva.numero_libros = estadistica_nueva.numero_libros + 1
            estadistica_nueva.numero_visitas_totales = estadistica_nueva.numero_visitas_totales + int(libro.visitas_mensuales)

        else: 
            nuevaEstadistica = EstadisticasPorMes(
            mes = libro.mes_creacion,
            anyo = libro.anyo_creacion,
            numero_libros = libros_totales + 1,
            numero_estimaciones = 0,
            numero_usuarios = 1,
            libro_mas_visitado = libro.titulo,
            isbn_libro_mas_visitado = libro.isbn,
            visitas_libro_mas_visitado = libro.visitas_mensuales,
            url_imagen_libro_mas_visitado = libro.url_imagen,
            numero_visitas_totales = libro.visitas_mensuales,
            numero_sugerencias = 0
            )
            db.session.add(nuevaEstadistica)
    db.session.query(EstadisticasPorMesAuxiliar).delete()
    db.session.commit()



class ExportarCSV(Resource):
    @jwt_required()
    def get(self):
        si = StringIO()
        cw = csv.writer(si, delimiter=';', quotechar='"', quoting=csv.QUOTE_MINIMAL)

      
        cw.writerow(['ID', 'Título', 'ISBN', 'Editorial', 'Descripción', 'Año de publicación', 'Puntuación', 'Ubicación del estudio', 'URL de la imagen', 'Visitas mensuales', 'Visitas totales', 'Mes de creación', 'Año de creación', 'Puntuación Lenguaje genérico', 'Puntuación Menores', 'Puntuación Adultos', 'Puntuación Ubicación', 'Puntuación Actividades'])

        libros = Libros.query.all()
        for libro in libros:
            libro.isbn = "\"" + libro.isbn + "\""
            cw.writerow([
                libro.id, libro.titulo,   libro.isbn, libro.editorial, libro.descripcion, libro.anyo_publicacion,
                libro.puntuacion, libro.ubicacion_estudio, libro.url_imagen, libro.visitas_mensuales, libro.visitas_totales,
                libro.mes_creacion, libro.anyo_creacion, libro.puntuacion_masculino_generico, libro.puntuacion_menores,
                libro.puntuacion_adultos, libro.puntuacion_ubicacion, libro.puntuacion_actividades
            ])

        output = si.getvalue()
        output = '\ufeff' + output  
        output = output.encode('utf-8')

        return Response(
            output,
            mimetype="text/csv",
            headers={"Content-Disposition": "attachment;filename=libros.csv"}
        )

class ExportarExcel(Resource):
    @jwt_required()
    def get(self):
        wb = Workbook()
        ws = wb.active

        
        ws.append(['ID', 'Título', 'ISBN', 'Editorial', 'Descripción', 'Año de publicación', 'Puntuación', 'Ubicación del estudio', 'URL de la imagen', 'Visitas mensuales', 'Visitas totales', 'Mes de creación', 'Año de creación', 'Puntuación Lenguaje genérico', 'Puntuación Menores', 'Puntuación Adultos', 'Puntuación Ubicación', 'Puntuación Actividades'])

        libros = Libros.query.all()
        for libro in libros:
            libro.isbn = "\"" + libro.isbn + "\""
            ws.append([
                libro.id, libro.titulo, libro.isbn, libro.editorial, libro.descripcion, libro.anyo_publicacion,
                libro.puntuacion, libro.ubicacion_estudio, libro.url_imagen, libro.visitas_mensuales, libro.visitas_totales,
                libro.mes_creacion, libro.anyo_creacion, libro.puntuacion_masculino_generico, libro.puntuacion_menores,
                libro.puntuacion_adultos, libro.puntuacion_ubicacion, libro.puntuacion_actividades
            ])

        si = BytesIO()
        wb.save(si)
        si.seek(0)

        return send_file(
            si,
            as_attachment=True,
            download_name="libros.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )




class ImportarArchivo(Resource):
    @jwt_required()
    def post(self):
        if 'archivo' not in request.files:
            return {"mensaje": "No se envió el archivo."}, 400

        archivo = request.files['archivo']
        filename = secure_filename(archivo.filename)
        ClonarEstadisticas()
        if filename.endswith('.csv'):
            return self.importar_csv(archivo)
        elif filename.endswith('.xlsx'):
            return self.importar_excel(archivo)
        else:
            return {"mensaje": "Formato de archivo no soportado."}, 400
    @jwt_required()
    def importar_csv(self, archivo):
        stream = StringIO(archivo.stream.read().decode("UTF-8"), newline=None)
        csv_input = csv.reader(stream, delimiter=';')

        Libros.query.delete()
        for i, row in enumerate(csv_input):
            if i == 0:  
                continue
            isbn = row[2].strip("'")
            isbn = isbn.strip('"')
            nuevo_libro = Libros(
                titulo=row[1],
                isbn=isbn,
                editorial=row[3],
                descripcion=row[4],
                anyo_publicacion=row[5],
                puntuacion=row[6] if row[6] else 0,
                ubicacion_estudio=row[7],
                url_imagen=row[8],
                visitas_mensuales=row[9] if row[9] else 0,
                visitas_totales=row[10] if row[10] else 0,
                mes_creacion=row[11],
                anyo_creacion=row[12],
                puntuacion_masculino_generico=int(row[13]),
                puntuacion_menores=int(row[14]),
                puntuacion_adultos=int(row[15]),
                puntuacion_ubicacion=int(row[16]),
                puntuacion_actividades=int(row[17])
            )
            db.session.add(nuevo_libro)
            ActualizarEstadisticasImportacion(nuevo_libro)
        
        fecha_modificacion.actualizar_fecha_modificacion()
        db.session.commit()

        return jsonify({"mensaje": "Datos importados exitosamente desde CSV"})

    @jwt_required()
    def importar_excel(self, archivo):
        wb = load_workbook(filename=BytesIO(archivo.read()))
        ws = wb.active

        Libros.query.delete()
        for row in ws.iter_rows(min_row=2):
            
            isbn = row[2].value.strip("'")
            isbn = isbn.strip('"')
            nuevo_libro = Libros(
                titulo=row[1].value,
                isbn=row[2].value,
                editorial=row[3].value,
                descripcion=row[4].value,
                anyo_publicacion=row[5].value,
                puntuacion=int(row[6].value) if row[6].value else None,
                ubicacion_estudio=row[7].value,
                url_imagen=row[8].value,
                visitas_mensuales=int(row[9].value) if row[9].value else 0,
                visitas_totales=int(row[10].value) if row[10].value else 0,
                mes_creacion=row[11].value or datetime.now().strftime("%m").lstrip("0"),
                anyo_creacion=row[12].value or datetime.now().strftime("%Y"),
                puntuacion_masculino_generico=int(row[13]),
                puntuacion_menores=int(row[14]),
                puntuacion_adultos=int(row[15]),
                puntuacion_ubicacion=int(row[16]),
                puntuacion_actividades=int(row[17])
            )
            db.session.add(nuevo_libro)
            ActualizarEstadisticasImportacion(nuevo_libro)

        fecha_modificacion.actualizar_fecha_modificacion()
        

        db.session.commit()


        return jsonify({"mensaje": "Datos importados exitosamente desde Excel"})
    
    
    
    
    
    